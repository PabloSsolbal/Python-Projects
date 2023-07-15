from openpyxl import load_workbook  # ? openpyxl to work with Excel
import pandas as pd  # ? pandas to work with Excel
import sqlite3 as sql  # ? sqlite3 to query the db
import os  # ? to manage the archives

# ? reportlab to work with the pdf
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors


# ? Create Excel File
# * This function creates an Excel file ('data.xlsx') if it doesn't already exist.
# * - It checks if the file exists using the `os.path.exists()` function.
# * - If the file doesn't exist, it connects to the SQLite database ('data.db') using `sql.connect()`.
# * - It retrieves data from the 'Users', 'Pets', and 'Categorys' tables using SQL queries.
# * - It uses `pd.ExcelWriter()` to create an Excel writer object.
# * - It saves the dataframes to separate sheets in the Excel file using `to_excel()`.
# * - The sheet names are set to 'Users', 'Pets', and 'Categories'.
# * - The resulting Excel file is saved as 'data.xlsx'.


def create_excel():

    if not os.path.exists('data.xlsx'):

        with sql.connect('data.db') as conn:

            users_df = pd.read_sql_query("SELECT * FROM Users", conn)
            pets_df = pd.read_sql_query("SELECT * FROM Pets", conn)
            categories_df = pd.read_sql_query("SELECT * FROM Categorys", conn)

            with pd.ExcelWriter('data.xlsx') as writer:

                users_df.to_excel(writer, sheet_name='Users', index=False)
                pets_df.to_excel(writer, sheet_name='Pets', index=False)
                categories_df.to_excel(
                    writer, sheet_name='Categories', index=False)
    return


# ? Create Database
# * This function creates a SQLite database file ('data.db') if it doesn't already exist.
# * - It checks if the file exists using the `os.path.exists()` function.
# * - If the file exists, it returns immediately.
# * - If the file doesn't exist, it creates the database and necessary tables.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It creates three tables: 'Users', 'Pets', and 'Categorys' using SQL `CREATE TABLE` statements.
# * - The 'Users' table has columns: 'UserID' (INTEGER), 'Name' (TEXT), 'Lastname' (TEXT), with 'UserID' as the primary key.
# * - The 'Pets' table has columns: 'PetID' (INTEGER), 'CategoryID' (INTEGER), 'Name' (TEXT), 'Sex' (TEXT), 'UserID' (INTEGER), 'Age' (INTEGER), with 'PetID' as the primary key.
# * - The 'Categorys' table has columns: 'CategoryID' (INTEGER), 'Name' (TEXT), with 'CategoryID' as the primary key.
# * - After creating the tables, it returns the string "ok" to indicate successful execution.

def create_db():

    if os.path.exists("data.db"):
        return

    else:

        db = os.path.abspath("data.db")

        with sql.connect(db) as con:
            cur = con.cursor()
            cur.execute("""
CREATE TABLE "Users" (
	"UserID"	INTEGER,
	"Name"	TEXT,
	"Lastname"	TEXT,
	PRIMARY KEY("UserID" AUTOINCREMENT)
);""")
            cur.execute("""
CREATE TABLE "Pets" (
	"PetID"	INTEGER,
	"CategoryID"	INTEGER,
	"Name"	TEXT,
	"Sex"	TEXT,
	"UserID"	INTEGER,
	"Age"	INTEGER,
	PRIMARY KEY("PetID" AUTOINCREMENT)
);
""")
            cur.execute("""
CREATE TABLE "Categorys" (
	"CategoryID"	INTEGER,
	"Name"	TEXT,
	PRIMARY KEY("CategoryID" AUTOINCREMENT)
);
""")
            return "ok"


# ? Get Users
# * This function retrieves all users from the 'Users' table in the database.
# * - It checks if the database file 'data.db' exists using `os.path.exists()` function.
# * - If the file exists, it establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to select all rows from the 'Users' table using `cur.execute()`.
# * - It fetches all the rows using `cur.fetchall()` and returns the result.
# * - If the database file doesn't exist, it prints a message "connection failed".
def get_users():
    if os.path.exists("data.db"):
        with sql.connect("data.db") as con:
            cur = con.cursor()
            cur.execute("SELECT * FROM Users")
            data = cur.fetchall()
            return data
    else:
        print("connection failed")


# ? New User
# * This function creates a new user in the 'Users' table of the database.
# ! @param Name - The user's name.
# ! @param Lastname - The user's lastname.
# * takes the parameters 'Name' and 'Lastname' to specify the user's name and lastname.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It creates a new user record in the 'Users' table by executing an SQL INSERT statement
# *   with the provided name and lastname values using `cur.execute()`.
# * - It commits the transaction to save the changes using `con.commit()`.

def new_user(Name, Lastname):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"INSERT INTO Users(Name, Lastname) VALUES ('{Name}','{Lastname}')")
        con.commit()


# ? Delete User
# * This function deletes a user from the 'Users' table of the database based on the provided ID.
# ! @param ID - The user's ID.
# * - It takes the parameter 'ID' to specify the user's ID.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It deletes the user record from the 'Users' table by executing an SQL DELETE statement
# *   with the specified ID using `cur.execute()`.

def delete_user(ID):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(f"DELETE FROM Users WHERE UserID={ID}")


# ? Update User
# * This function updates the information of a user in the 'Users' table of the database based on the provided ID.
# ! @param ID - The user's ID.
# * - It takes the parameters 'ID', 'Name', and 'Lastname' to specify the user's ID and the updated name and lastname.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It updates the user record in the 'Users' table by executing an SQL UPDATE statement
# *   with the specified ID, name, and lastname using `cur.execute()`.

def update_user(ID, Name, Lastname):
    with sql.connect("data.db") as con:
        cur = con.cursor()
        cur.execute(
            f"UPDATE Users SET Name='{Name}', Lastname='{Lastname}' WHERE UserID={ID}")


def user_pets(userID):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"""
            SELECT p.PetID, c.Name AS Category, p.Name, p.Sex, p.Age FROM Pets AS p
            Join Categorys As c On p.CategoryID = c.CategoryID
            WHERE p.UserID = {userID}
            """)

        if cur.rowcount == 0:
            return None

        data = cur.fetchall()
        return data


# ? Get Pets
# * This function retrieves the information of all pets from the database.
# * - It checks if the database file 'data.db' exists using `os.path.exists()`.
# * - If the file exists, it establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to fetch the pet records from the 'Pets' table, along with additional
# *   information from the 'Categorys' and 'Users' tables using JOIN operations.
# * - It returns the fetched data as a result.
# * - If the database file does not exist, it prints a message indicating the connection failure.

def get_pets():

    if os.path.exists("data.db"):

        with sql.connect("data.db") as con:

            cur = con.cursor()
            cur.execute(
                """
                SELECT p.PetID, c.Name AS Category, p.Name, p.Sex,u.Name AS Owner, p.Age FROM Pets AS p
                Join Categorys As c On p.CategoryID = c.CategoryID
                Join Users AS u On p.UserID = u.UserID
                """
            )

            data = cur.fetchall()
            return data
    else:
        print("connection failed")


# ? Get Pet by ID
# * This function retrieves the information of a specific pet from the database based on the provided ID.
# ! @param ID - The pet's ID.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to fetch the pet record with the specified ID from the 'Pets' table.
# * - If no pet is found with the given ID (indicated by `cur.rowcount == 0`), it returns None.
# * - If a pet is found, it retrieves the fetched data as a result.
# * Note: The provided ID should be a unique identifier for a pet.

def get_pet(id):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"SELECT * FROM Pets WHERE PetID='{id}'")

        if cur.rowcount == 0:
            return None

        data = cur.fetchone()
        return data


# ? Update Pet Information
# * This function updates the information of a pet in the database based on the provided parameters.
# !@param petID - The pet's ID.
# !@param Category - The pet's category.
# !@param Name - The pet's name.
# !@param Sex - sex of the pet
# !@param Owner - The pet's owner.
# !@param Age - pet's age.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to update the pet record with the specified ID.
# * - The parameters `petID`, `Category`, `Name`, `Sex`, `Owner`, and `Age` are used to update the corresponding columns of the pet record.

def update_pet(petID, Category, Name, Sex, Owner, Age):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"UPDATE Pets Set CategoryID={Category}, Name='{Name}', Sex='{Sex}',UserID={Owner}, Age={Age} WHERE PetID='{petID}'")


# ? Create New Pet
# * This function creates a new pet in the database with the provided information.
# !@param userID - The user ID.
# !@param CategoryID - The pet's category.
# !@param Name - The pet's name.
# !@param Sex - sex of the pet
# !@param Age - pet's age.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to insert a new pet record with the specified `userID`, `categoryID`, `name`, `sex`, and `age`.
# * - The new pet record is committed to the database using `con.commit()`.

def create_pet(userID, categoryID, name, sex, age):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"INSERT INTO Pets (UserID, CategoryID, Name, Sex, Age) VALUES ({userID}, {categoryID}, '{name}', '{sex}', {age})")
        con.commit()


# Delete Pet
# * This function deletes a pet from the database based on the provided ID.
# !@param ID - The pet's ID.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to delete the pet record with the specified ID from the 'Pets' table.

def delete_pet(id):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"DELETE FROM Pets WHERE PetID={id}")


# ? View Categories
# * This function retrieves all the categories from the database and returns the category data.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to fetch all the category records from the 'Categorys' table.
# * - The fetched data is returned as a result.

def view_categorys():

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            """
            SELECT * FROM Categorys
            """
        )
        data = cur.fetchall()
        return data


# ? Create New Category
# * This function creates a new category in the database with the provided name.
# !@param name - The category's name.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to insert a new category record with the specified name.
# * - The new category record is committed to the database using `con.commit()`.
def create_category(name):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"INSERT INTO Categorys(Name) VALUES ('{name}')")
        con.commit()


# ? Delete Category
# * This function deletes a category from the database based on the provided ID.
# !@param ID - The category's ID.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to delete the category record with the specified ID from the 'Categorys' table.

def delete_category(id):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"DELETE FROM Categorys WHERE CategoryID={id}")


# ? Get Pets by Category
# * This function retrieves the pets belonging to a specific category from the database based on the provided category ID.
# !@param categoryID - The category's ID.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to fetch the pets with the specified category ID from the 'Pets' table, along with their owners' information.
# * - If no pets are found with the given category ID (indicated by `cur.rowcount == 0`), it returns None.
# * - If pets are found, the fetched data is returned as a result.

def category_pets(categoryID):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"""SELECT p.PetID, p.Name, p.Sex,u.Name AS Owner, p.Age FROM Pets AS p
            Join Users AS u On p.UserID = u.UserID
            WHERE p.CategoryID={categoryID}
            """)

        if cur.rowcount == 0:
            return None

        data = cur.fetchall()
        return data


# ? Get Category by ID
# * This function retrieves the information of a specific category from the database based on the provided ID.
# !@param id - The category's ID.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It executes an SQL query to fetch the category name with the specified ID from the 'Categorys' table.
# * - If no category is found with the given ID (indicated by `cur.rowcount == 0`), it returns None.
# * - If a category is found, it retrieves the fetched data as a result.

def get_category(id):

    with sql.connect("data.db") as con:

        cur = con.cursor()
        cur.execute(
            f"SELECT Name FROM Categorys WHERE CategoryID={id}")

        if cur.rowcount == 0:
            return None

        data = cur.fetchone()
        return data


# ?Export Data to Excel
# * This function exports the data from the database tables to an Excel file.
# * - It calls the `create_excel()` function to generate the Excel file with the data from the database.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It loads the existing Excel file using `load_workbook()`.
# * - It creates a new Excel writer using `pd.ExcelWriter()` and attaches it to the loaded workbook.
# * - It reads the data from the 'Users', 'Pets', and 'Categorys' tables into dataframes.
# * - It writes the dataframes to separate sheets in the Excel file using `to_excel()`.
# * - The changes are saved to the Excel file using `writer.save()`.

def export_excel():

    create_excel()

    with sql.connect('data.db') as conn:

        book = load_workbook('data.xlsx')

        writer = pd.ExcelWriter('data.xlsx', engine='openpyxl')
        writer.book = book

        users_df = pd.read_sql_query("SELECT * FROM Users", conn)
        users_df.to_excel(writer, sheet_name='Users', index=False)

        pets_df = pd.read_sql_query("SELECT * FROM Pets", conn)
        pets_df.to_excel(writer, sheet_name='Pets', index=False)

        categories_df = pd.read_sql_query("SELECT * FROM Categorys", conn)
        categories_df.to_excel(writer, sheet_name='Categories', index=False)

        writer.save()


# ? Export Data to PDF
# * This function exports the data from the database tables to a PDF file.
# * - It establishes a connection to the database using `sql.connect()`.
# * - It reads the data from the 'Users', 'Pets', and 'Categorys' tables into dataframes.
# * - It creates a new PDF document using `SimpleDocTemplate`.
# * - It creates a list to store the tables for each dataframe.
# * - For each dataframe, it creates a table object and customizes its style using `TableStyle`.
# * - The tables are added to the table list.
# * - The table list is used to build the PDF document using `pdf.build()`.

def export_pdf():

    with sql.connect('data.db') as conn:

        users = pd.read_sql_query("SELECT * FROM Users", conn)

        pets = pd.read_sql_query("""
                SELECT p.PetID, c.Name AS Category, p.Name, p.Sex,u.Name AS Owner, p.Age FROM Pets AS p
                Join Categorys As c On p.CategoryID = c.CategoryID
                Join Users AS u On p.UserID = u.UserID
                """, conn)

        categorys = pd.read_sql_query("SELECT * FROM Categorys", conn)

        dataframes = [users, pets, categorys]

        pdf_filename = "tables.pdf"

        pdf = SimpleDocTemplate(pdf_filename, pagesize=letter)

        table_list = []

        for df in dataframes:

            table_data = [df.columns.tolist()]+df.values.tolist()
            table = Table(table_data)

            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.whitesmoke),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))

            table_list.append(table)
        pdf.build(table_list)
